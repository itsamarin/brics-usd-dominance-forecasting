"""
Add Charts to Predictive Analysis Forecasts
This script adds professional charts to the forecast Excel workbook.

Run this after generating the main forecast file:
    python predictive_analysis_forecast.py
    python add_charts_to_forecasts.py
"""

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
import sys


def add_btc_chart(ws):
    """Add chart to BTC forecast sheet."""
    # Find last row with data
    last_row = 6
    while ws.cell(row=last_row, column=1).value is not None:
        last_row += 1
    last_row -= 1
    
    # Create line chart
    chart = LineChart()
    chart.title = "Bitcoin USD Trading Volume - 3-Month MA Forecast"
    chart.style = 13
    chart.y_axis.title = 'Trading Volume (BTC)'
    chart.x_axis.title = 'Date'
    chart.height = 10
    chart.width = 20
    
    # Add actual volume data
    data_ref = Reference(ws, min_col=3, min_row=5, max_row=last_row)
    dates_ref = Reference(ws, min_col=2, min_row=6, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(dates_ref)
    
    # Add 3-month MA
    ma_ref = Reference(ws, min_col=4, min_row=5, max_row=last_row)
    chart.add_data(ma_ref, titles_from_data=True)
    
    # Add forecast
    forecast_ref = Reference(ws, min_col=5, min_row=5, max_row=last_row)
    chart.add_data(forecast_ref, titles_from_data=True)
    
    # Style the series
    chart.series[0].graphicalProperties.line.solidFill = "4472C4"  # Blue
    chart.series[0].graphicalProperties.line.width = 2.25
    chart.series[1].graphicalProperties.line.solidFill = "70AD47"  # Green
    chart.series[1].graphicalProperties.line.width = 2.25
    if len(chart.series) > 2:
        chart.series[2].graphicalProperties.line.solidFill = "FF0000"  # Red
        chart.series[2].graphicalProperties.line.width = 3
        chart.series[2].graphicalProperties.line.dashStyle = "dash"
    
    # Add chart to sheet
    ws.add_chart(chart, "H5")
    print("  ✓ BTC chart added")


def add_gold_charts(ws):
    """Add charts to Gold BRICS forecast sheet."""
    # Find last row with data
    last_row = 6
    while ws.cell(row=last_row, column=1).value is not None:
        last_row += 1
    last_row -= 1
    
    # Chart 1: Quantity
    chart_qty = LineChart()
    chart_qty.title = "BRICS Gold Imports (Quantity) - 3-Month MA Forecast"
    chart_qty.style = 13
    chart_qty.y_axis.title = 'Quantity (kg)'
    chart_qty.x_axis.title = 'Date'
    chart_qty.height = 10
    chart_qty.width = 20
    
    data_ref = Reference(ws, min_col=3, min_row=5, max_row=last_row)
    dates_ref = Reference(ws, min_col=2, min_row=6, max_row=last_row)
    chart_qty.add_data(data_ref, titles_from_data=True)
    chart_qty.set_categories(dates_ref)
    
    ma_ref = Reference(ws, min_col=5, min_row=5, max_row=last_row)
    chart_qty.add_data(ma_ref, titles_from_data=True)
    
    chart_qty.series[0].graphicalProperties.line.solidFill = "FFC000"  # Gold
    chart_qty.series[0].graphicalProperties.line.width = 2.25
    chart_qty.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red
    chart_qty.series[1].graphicalProperties.line.width = 3
    chart_qty.series[1].graphicalProperties.line.dashStyle = "dash"
    
    ws.add_chart(chart_qty, "I5")
    
    # Chart 2: Value
    chart_val = LineChart()
    chart_val.title = "BRICS Gold Imports (Value USD) - 3-Month MA Forecast"
    chart_val.style = 13
    chart_val.y_axis.title = 'Value (USD)'
    chart_val.x_axis.title = 'Date'
    chart_val.height = 10
    chart_val.width = 20
    
    data_ref = Reference(ws, min_col=4, min_row=5, max_row=last_row)
    chart_val.add_data(data_ref, titles_from_data=True)
    chart_val.set_categories(dates_ref)
    
    ma_ref = Reference(ws, min_col=6, min_row=5, max_row=last_row)
    chart_val.add_data(ma_ref, titles_from_data=True)
    
    chart_val.series[0].graphicalProperties.line.solidFill = "FFC000"  # Gold
    chart_val.series[0].graphicalProperties.line.width = 2.25
    chart_val.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red
    chart_val.series[1].graphicalProperties.line.width = 3
    chart_val.series[1].graphicalProperties.line.dashStyle = "dash"
    
    ws.add_chart(chart_val, "I25")
    
    print("  ✓ Gold charts added (Quantity & Value)")


def add_oil_charts(ws):
    """Add charts to Oil BRICS forecast sheet."""
    # Find last row with data
    last_row = 6
    while ws.cell(row=last_row, column=1).value is not None:
        last_row += 1
    last_row -= 1
    
    # Chart 1: Quantity
    chart_qty = LineChart()
    chart_qty.title = "BRICS Crude Oil Imports (Quantity) - 3-Month MA Forecast"
    chart_qty.style = 13
    chart_qty.y_axis.title = 'Quantity (kg)'
    chart_qty.x_axis.title = 'Date'
    chart_qty.height = 10
    chart_qty.width = 20
    
    data_ref = Reference(ws, min_col=3, min_row=5, max_row=last_row)
    dates_ref = Reference(ws, min_col=2, min_row=6, max_row=last_row)
    chart_qty.add_data(data_ref, titles_from_data=True)
    chart_qty.set_categories(dates_ref)
    
    ma_ref = Reference(ws, min_col=5, min_row=5, max_row=last_row)
    chart_qty.add_data(ma_ref, titles_from_data=True)
    
    chart_qty.series[0].graphicalProperties.line.solidFill = "000000"  # Black
    chart_qty.series[0].graphicalProperties.line.width = 2.25
    chart_qty.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red
    chart_qty.series[1].graphicalProperties.line.width = 3
    chart_qty.series[1].graphicalProperties.line.dashStyle = "dash"
    
    ws.add_chart(chart_qty, "I5")
    
    # Chart 2: Value
    chart_val = LineChart()
    chart_val.title = "BRICS Crude Oil Imports (Value USD) - 3-Month MA Forecast"
    chart_val.style = 13
    chart_val.y_axis.title = 'Value (USD)'
    chart_val.x_axis.title = 'Date'
    chart_val.height = 10
    chart_val.width = 20
    
    data_ref = Reference(ws, min_col=4, min_row=5, max_row=last_row)
    chart_val.add_data(data_ref, titles_from_data=True)
    chart_val.set_categories(dates_ref)
    
    ma_ref = Reference(ws, min_col=6, min_row=5, max_row=last_row)
    chart_val.add_data(ma_ref, titles_from_data=True)
    
    chart_val.series[0].graphicalProperties.line.solidFill = "000000"  # Black
    chart_val.series[0].graphicalProperties.line.width = 2.25
    chart_val.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red
    chart_val.series[1].graphicalProperties.line.width = 3
    chart_val.series[1].graphicalProperties.line.dashStyle = "dash"
    
    ws.add_chart(chart_val, "I25")
    
    print("  ✓ Oil charts added (Quantity & Value)")


def main(input_file='Predictive_Analysis_Forecasts.xlsx', 
         output_file='Predictive_Analysis_Forecasts_with_Charts.xlsx'):
    """
    Add charts to the forecast workbook.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file (with charts)
    """
    print("="*70)
    print("Adding Charts to Predictive Analysis Forecasts")
    print("="*70)
    
    print(f"\n[1/4] Loading workbook: {input_file}")
    try:
        wb = load_workbook(input_file)
    except FileNotFoundError:
        print(f"\n❌ Error: File '{input_file}' not found!")
        print("Please run 'python predictive_analysis_forecast.py' first.")
        sys.exit(1)
    
    print("\n[2/4] Adding charts to forecast sheets...")
    
    # Add BTC chart
    if 'BTC_Forecast' in wb.sheetnames:
        add_btc_chart(wb['BTC_Forecast'])
    else:
        print("  ⚠ BTC_Forecast sheet not found, skipping")
    
    # Add Gold charts
    if 'Gold_BRICS_Forecast' in wb.sheetnames:
        add_gold_charts(wb['Gold_BRICS_Forecast'])
    else:
        print("  ⚠ Gold_BRICS_Forecast sheet not found, skipping")
    
    # Add Oil charts
    if 'Oil_BRICS_Forecast' in wb.sheetnames:
        add_oil_charts(wb['Oil_BRICS_Forecast'])
    else:
        print("  ⚠ Oil_BRICS_Forecast sheet not found, skipping")
    
    print("\n[3/4] Saving workbook with charts...")
    wb.save(output_file)
    print(f"  ✓ Saved: {output_file}")
    
    print("\n[4/4] Complete!")
    print("="*70)
    print("SUCCESS! Charts added to forecast workbook.")
    print("="*70)
    print(f"\nOutput file: {output_file}")
    print("\nCharts included:")
    print("  • BTC_Forecast: 1 line chart (Actual, MA, Forecast)")
    print("  • Gold_BRICS_Forecast: 2 line charts (Quantity & Value)")
    print("  • Oil_BRICS_Forecast: 2 line charts (Quantity & Value)")
    print("\nTotal: 5 professional charts with forecast visualization")
    print("="*70)


if __name__ == '__main__':
    # Check for command line arguments
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else \
                     input_file.replace('.xlsx', '_with_Charts.xlsx')
        main(input_file, output_file)
    else:
        main()
